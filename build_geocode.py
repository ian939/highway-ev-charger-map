# 휴게소 충전소 엑셀을 읽어 Kakao 지오코딩 후 data.json을 생성하는 빌드 스크립트
# -*- coding: utf-8 -*-
import os
import re
import csv
import json
import sys
import io
import time

import requests
import openpyxl
from dotenv import load_dotenv

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "고속도로 휴게소 전기차충전소 현황_260619.xlsx")
OUT_JSON = os.path.join(BASE, "data.json")
FAILED_CSV = os.path.join(BASE, "geocode_failed.csv")

load_dotenv(os.path.join(BASE, ".env"))
REST_KEY = os.environ["KAKAO_REST_KEY"]
HEADERS = {"Authorization": "KakaoAK " + REST_KEY}

# 운영사 컬럼 정의 (엑셀 F~V = 인덱스 6~22). 카테고리별 그룹.
OPERATOR_COLS = [
    (6, "환경부", "환경부"),
    (7, "한전", "한전"),
    (8, "현대 E-pit", "기타"),
    (9, "시그넷", "기타"),
    (10, "SC생활안전", "기타"),
    (11, "대영채비", "기타"),
    (12, "1차 SK일렉링크", "SK·차수"),
    (13, "2차 이지차저", "SK·차수"),
    (14, "3차 SK시그넷", "SK·차수"),
    (15, "3차 워터", "SK·차수"),
    (16, "4차 SK일렉링크", "SK·차수"),
    (17, "4차 워터", "SK·차수"),
    (18, "4차 채비", "SK·차수"),
    (19, "5차 1권역", "SK·차수"),
    (20, "5차 2권역", "SK·차수"),
    (21, "5차 3권역", "SK·차수"),
    (22, "5차 멀티(50kW)", "SK·차수"),
]

MAIN_LAST_ROW = 237  # 메인 섹션 마지막 행(번호 236). 이후는 보조 섹션 → 제외.

KR_LAT = (33.0, 39.5)
KR_LNG = (124.0, 132.0)


def in_korea(lat, lng):
    return KR_LAT[0] <= lat <= KR_LAT[1] and KR_LNG[0] <= lng <= KR_LNG[1]


def extract_road_no(address):
    # 주소에서 '○○고속도로 232' 형태의 도로 번호 추출
    m = re.search(r"고속도로\s*(\d+)", address or "")
    return m.group(1) if m else None


def kakao_keyword(query):
    r = requests.get(
        "https://dapi.kakao.com/v2/local/search/keyword.json",
        headers=HEADERS,
        params={"query": query, "size": 10},
        timeout=10,
    )
    r.raise_for_status()
    return r.json().get("documents", [])


def kakao_address(query):
    r = requests.get(
        "https://dapi.kakao.com/v2/local/search/address.json",
        headers=HEADERS,
        params={"query": query, "size": 5},
        timeout=10,
    )
    r.raise_for_status()
    return r.json().get("documents", [])


def clean_address_for_search(address):
    # 주소 끝에 공백 없이 붙은 '…휴게소' 시설명을 제거하고 도로명+번지까지만 남김
    addr = address or ""
    # '… 고속도로 232강릉(강릉)휴게소' → '… 고속도로 232'
    m = re.search(r"^(.*?고속도로\s*\d+(?:-\d+)?)", addr)
    if m:
        return m.group(1).strip()
    # 일반 도로명: '…로 1089경산(서울)휴게소' → '…로 1089'
    m = re.search(r"^(.*?[로길]\s*\d+(?:-\d+)?)", addr)
    if m:
        return m.group(1).strip()
    return addr.strip()


def pick_best(docs, direction, road_no):
    """고속도로휴게소 우선 + 방향/도로번호 매칭으로 최적 결과 선택."""
    rest = [d for d in docs if "고속도로휴게소" in (d.get("category_name") or "")]
    pool = rest or [d for d in docs if "휴게소" in (d.get("category_name") or "")]
    if not pool:
        pool = docs
    if not pool:
        return None
    # 1) 방향 매칭 (예: 부산방향)
    if direction:
        for d in pool:
            if f"{direction}방향" in (d.get("place_name") or ""):
                return d
    # 2) 도로번호 매칭 (road_address_name 끝 번호)
    if road_no:
        for d in pool:
            ra = d.get("road_address_name") or d.get("address_name") or ""
            if re.search(r"\b" + re.escape(road_no) + r"\b", ra):
                return d
    # 3) 첫 결과
    return pool[0]


def geocode(name, address):
    # name: '강릉(강릉)' → base '강릉', direction '강릉'
    base = re.split(r"[(\s]", name.strip())[0]
    mdir = re.search(r"\(([^)]+)\)", name)
    direction = mdir.group(1) if mdir else None
    road_no = extract_road_no(address)

    queries = [base + "휴게소", base + " 휴게소"]
    if base.endswith("휴게소"):
        queries = [base, base[:-3] + "휴게소"]
    for q in queries:
        try:
            docs = kakao_keyword(q)
        except Exception as e:
            print("  키워드 검색 오류:", q, e)
            docs = []
        best = pick_best(docs, direction, road_no)
        if best:
            lat, lng = float(best["y"]), float(best["x"])
            if in_korea(lat, lng):
                return lat, lng, "keyword:" + q, best.get("place_name", "")

    # 폴백: 정제 주소로 주소검색
    cq = clean_address_for_search(address)
    if cq:
        try:
            docs = kakao_address(cq)
        except Exception as e:
            print("  주소 검색 오류:", cq, e)
            docs = []
        if docs:
            d = docs[0]
            lat, lng = float(d["y"]), float(d["x"])
            if in_korea(lat, lng):
                return lat, lng, "address:" + cq, d.get("address_name", "")

    # 폴백2: 건설중/미확정 부지 → 읍·면·리 단위 근사 좌표 (approx)
    region = re.sub(r"\(건설중\)|건설구간|일원|\s+산?\s*\d+(-\d+)?$", "", address or "").strip()
    region = re.sub(r"\([^)]*\)", "", region).strip()
    if region:
        try:
            docs = kakao_address(region)
        except Exception:
            docs = []
        if docs:
            d = docs[0]
            lat, lng = float(d["y"]), float(d["x"])
            if in_korea(lat, lng):
                return lat, lng, "approx:" + region, d.get("address_name", "")
    return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    items = []
    failed = []
    skipped_no_op = 0

    for r in range(2, MAIN_LAST_ROW + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        ftype = ws.cell(r, 3).value
        route = ws.cell(r, 4).value
        address = ws.cell(r, 5).value

        operators = {}
        total = 0
        for col, label, _cat in OPERATOR_COLS:
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)) and v > 0:
                operators[label] = int(v)
                total += int(v)
        if total == 0:
            skipped_no_op += 1
            continue  # 운영사 데이터 없는 휴게소 제외

        geo = geocode(str(name), str(address) if address else "")
        if geo is None:
            failed.append((r, name, address))
            print(f"  [실패] 행{r} {name} | {address}")
            continue
        lat, lng, method, matched = geo
        items.append({
            "no": ws.cell(r, 1).value,
            "name": name,
            "type": ftype,
            "route": route,
            "address": address,
            "lat": lat,
            "lng": lng,
            "operators": operators,
            "total": total,
            "approx": method.startswith("approx:"),
            "_method": method,
            "_matched": matched,
        })
        time.sleep(0.03)

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=1)

    if failed:
        with open(FAILED_CSV, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["row", "name", "address"])
            w.writerows(failed)

    total_chargers = sum(it["total"] for it in items)
    print("=" * 50)
    print(f"운영사 데이터 보유 휴게소(대상): {len(items) + len(failed)}곳")
    print(f"  지오코딩 성공: {len(items)}곳")
    print(f"  지오코딩 실패: {len(failed)}곳  (→ geocode_failed.csv)")
    print(f"운영사 데이터 없어 제외: {skipped_no_op}곳")
    print(f"총 충전기 수(성공분): {total_chargers}대")
    print(f"저장: {OUT_JSON}")


if __name__ == "__main__":
    main()
