# 휴게소 충전기 정보.xlsx → 휴게소별 용량 집계 JSON 생성 + data.json 좌표 매칭 리포트
import io, sys, json, os, math
from collections import defaultdict, Counter
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
HERE = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(HERE)
XLSX = os.path.join(PARENT, "휴게소 충전기 정보.xlsx")
OUT = os.path.join(HERE, "sk_chargers.json")

wb = load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["Sheet1"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
I = {h: i for i, h in enumerate(hdr)}
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]

def clean(s):
    # 고립 서로게이트 등 잘못된 유니코드 제거 (JSON/브라우저 호환)
    if isinstance(s, str):
        return s.encode("utf-8", "ignore").decode("utf-8")
    return s

def g(r, name):
    return clean(r[I[name]])

# 휴게소명 기준 그룹핑 (1행 = 충전기 1대)
groups = defaultdict(list)
for r in rows:
    groups[g(r, "충전소명")].append(r)

def speed_label(code):
    return "완속" if str(code) == "01" else "급속"

stations = []
for name, rs in groups.items():
    lats = [float(g(r, "GPS 위도")) for r in rs if g(r, "GPS 위도")]
    lngs = [float(g(r, "GPS 경도")) for r in rs if g(r, "GPS 경도")]
    lat = sum(lats) / len(lats)
    lng = sum(lngs) / len(lngs)
    # (kW, 급속/완속)별 대수 집계
    bykw = Counter()
    for r in rs:
        kw = int(float(g(r, "전력량")))
        bykw[(kw, speed_label(g(r, "01-완속, 02-급속")))] += 1
    by = [{"kw": kw, "speed": sp, "count": c}
          for (kw, sp), c in sorted(bykw.items(), key=lambda x: (-x[0][0], x[0][1]))]
    fast = sum(c for (kw, sp), c in bykw.items() if sp == "급속")
    slow = sum(c for (kw, sp), c in bykw.items() if sp == "완속")
    addr = next((g(r, "도로명 주소") for r in rs if g(r, "도로명 주소")), "") \
        or next((g(r, "지번 주소") for r in rs if g(r, "지번 주소")), "")
    conns = sorted({g(r, "커넥터") for r in rs if g(r, "커넥터")})
    makers = sorted({g(r, "제조사") for r in rs if g(r, "제조사")})
    years = sorted({str(g(r, "설치년도")) for r in rs if g(r, "설치년도")})
    lucky = sum(1 for r in rs if str(g(r, "럭키패스 적용여부")) == "Y")
    status = dict(Counter(str(g(r, "현장상태")) for r in rs))
    stations.append({
        "name": name.strip(),
        "lat": round(lat, 7), "lng": round(lng, 7),
        "address": (addr or "").strip(),
        "region": (g(rs[0], "지역") or "").strip(),
        "total": len(rs), "fast": fast, "slow": slow,
        "byKw": by, "connectors": conns, "makers": makers,
        "years": years, "luckypass": lucky, "status": status,
    })

stations.sort(key=lambda s: s["name"])
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(stations, f, ensure_ascii=False, indent=1)

# ---- 리포트 ----
print(f"[생성] {OUT}")
print(f"휴게소 {len(stations)}곳 / 충전기 {sum(s['total'] for s in stations)}대")
kw_total = Counter()
for s in stations:
    for b in s["byKw"]:
        kw_total[(b['kw'], b['speed'])] += b['count']
print("용량별 합계:", {f"{k}kW {sp}": c for (k, sp), c in sorted(kw_total.items(), key=lambda x: -x[0][0])})

# data.json 좌표 매칭 (가장 가까운 SK 휴게소까지 거리)
R = 6371
def hav(a, b):
    dl = math.radians(b[0]-a[0]); dn = math.radians(b[1]-a[1])
    s = math.sin(dl/2)**2 + math.cos(math.radians(a[0]))*math.cos(math.radians(b[0]))*math.sin(dn/2)**2
    return 2*R*math.asin(math.sqrt(s))

dj = json.load(open(os.path.join(PARENT, "data.json"), encoding="utf-8"))
dj_sk = [d for d in dj if any("SK일렉링크" in k for k in d.get("operators", {}))
         and isinstance(d.get("lat"), (int, float))]
print(f"\ndata.json SK 휴게소: {len(dj_sk)}곳")
matched = 0; far = []
for s in stations:
    best = min(((hav((s['lat'], s['lng']), (d['lat'], d['lng'])), d['name']) for d in dj_sk),
               default=(9e9, None))
    if best[0] <= 0.5:
        matched += 1
    else:
        far.append((s['name'], round(best[0], 1), best[1]))
print(f"좌표 0.5km 내 매칭: {matched}/{len(stations)}곳")
if far:
    print(f"미매칭(0.5km 초과) {len(far)}곳:")
    for nm, dist, near in far[:30]:
        print(f"  {nm}  → 최근접 {near} {dist}km")
